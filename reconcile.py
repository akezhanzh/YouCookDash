"""
reconcile.py — сверка акта сверки от поставщика с нашей БД.

Поддерживает:
  - XLSX акт (1С, формат «Реализация ТМЗ и услуг» в одной ячейке)
  - PDF акт (1С, текст разбит на 3 строки: номер / дата-сумма / от ДАТА)

Возвращает список накладных из акта + сравнивает с БД → отчёт
о расхождениях: чего нет в базе, что лишнее, где суммы разъехались.
"""
import re
import sqlite3
from pathlib import Path
from datetime import datetime

import pdfplumber
import openpyxl

from parse_invoice import clean_supplier_name, to_float


# ── Парсеры ────────────────────────────────────────────────────────────────────

RE_INV_INLINE = re.compile(
    r'Реализация\s+ТМЗ\s+и\s+услуг\s+(\d+)\s+от\s+(\d{2}\.\d{2}\.\d{4})',
    re.IGNORECASE,
)


def _header_supplier(text: str) -> str | None:
    """Из шапки акта («... между Индивидуальный предприниматель "МИЛАНА" и ТОО ...»)."""
    m = re.search(
        r'между\s+(.+?)\s+и\s+(ТОО|ИП|АО|ООО|Товарищество|Акционерное|Индивидуальный|Общество)',
        text, re.IGNORECASE | re.DOTALL,
    )
    if m:
        return clean_supplier_name(m.group(1).strip())
    # Fallback: «По данным <Поставщик>, KZT»
    m = re.search(r'По\s+данным\s+(.+?)(?:,\s*KZT|\s+KZT|$)', text, re.IGNORECASE)
    if m:
        return clean_supplier_name(m.group(1).strip())
    return None


def _header_period(text: str) -> tuple[str | None, str | None]:
    m = re.search(r'с\s+(\d{2}\.\d{2}\.\d{4})\s+по\s+(\d{2}\.\d{2}\.\d{4})', text)
    if m:
        def iso(d):
            dd, mm, yy = d.split('.')
            return f'{yy}-{mm}-{dd}'
        return iso(m.group(1)), iso(m.group(2))
    return None, None


def parse_akt_xlsx(path: Path) -> dict:
    """Парсит XLSX-акт 1С. Возвращает {supplier, period_from, period_to, invoices}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # Собираем весь текст шапки (первые 10 строк)
    header_text = ' '.join(
        str(c) for r in rows[:10] for c in r if c
    )
    supplier = _header_supplier(header_text)
    pfrom, pto = _header_period(header_text)

    invoices = []
    for row in rows:
        for c_idx, cell in enumerate(row):
            if not cell:
                continue
            m = RE_INV_INLINE.search(str(cell))
            if not m:
                continue
            inv_id = str(int(m.group(1)))  # strip leading zeros
            date_iso = _to_iso(m.group(2))
            # Сумма — в этой же строке, первая числовая колонка справа от документа
            amount = _extract_amount_from_row(row, c_idx)
            invoices.append({
                'invoice_id': inv_id,
                'date': date_iso,
                'amount': amount,
                'raw': str(cell).strip()[:80],
            })

    return {
        'supplier': supplier, 'period_from': pfrom, 'period_to': pto,
        'invoices': invoices, 'source': 'xlsx',
    }


def parse_akt_pdf(path: Path) -> dict:
    """Парсит PDF-акт. Строки типа «Реализация ТМЗ и услуг 54 / 03.03.26 323 974,00 / от 03.03.2026»."""
    all_lines = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            all_lines.extend(text.split('\n'))

    header_text = ' '.join(all_lines[:8])
    supplier = _header_supplier(header_text)
    pfrom, pto = _header_period(header_text)

    invoices = []
    i = 0
    while i < len(all_lines):
        line = all_lines[i]
        # Шаг 1: «Реализация ТМЗ и услуг N ...»
        m_num = re.search(r'Реализация\s+ТМЗ\s+и\s+услуг\s+(\d+)', line, re.IGNORECASE)
        if m_num:
            inv_id = str(int(m_num.group(1)))
            # Шаг 2: следующая строка содержит «DD.MM.YY <amount>»
            date_iso = None; amount = None
            if i + 1 < len(all_lines):
                m_amt = re.search(r'(\d{2})\.(\d{2})\.(\d{2})\s+([\d\s]+,\d{2})', all_lines[i + 1])
                if m_amt:
                    dd, mm, yy = m_amt.group(1), m_amt.group(2), m_amt.group(3)
                    date_iso = f'20{yy}-{mm}-{dd}'
                    amount = to_float(m_amt.group(4))
            # Шаг 3: подстрахуемся на строке «от DD.MM.YYYY»
            if i + 2 < len(all_lines):
                m_full = re.search(r'от\s+(\d{2}\.\d{2}\.\d{4})', all_lines[i + 2])
                if m_full and not date_iso:
                    date_iso = _to_iso(m_full.group(1))
            if date_iso and amount is not None:
                invoices.append({
                    'invoice_id': inv_id, 'date': date_iso, 'amount': amount,
                    'raw': line.strip()[:80],
                })
            i += 2
        i += 1

    return {
        'supplier': supplier, 'period_from': pfrom, 'period_to': pto,
        'invoices': invoices, 'source': 'pdf',
    }


def parse_akt(path: Path) -> dict:
    """Автоопределение XLSX vs PDF."""
    ext = Path(path).suffix.lower()
    if ext in ('.xlsx', '.xls'):
        return parse_akt_xlsx(Path(path))
    if ext == '.pdf':
        return parse_akt_pdf(Path(path))
    raise ValueError(f'Неподдерживаемый формат: {ext}')


def is_akt(path: Path) -> bool:
    """Проверяет по первым строкам, является ли файл актом сверки."""
    try:
        ext = Path(path).suffix.lower()
        if ext in ('.xlsx', '.xls'):
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > 4: break
                for cell in row:
                    if cell and 'Акт сверки' in str(cell):
                        return True
        elif ext == '.pdf':
            with pdfplumber.open(path) as pdf:
                if not pdf.pages: return False
                first = pdf.pages[0].extract_text() or ''
                return 'Акт сверки' in first[:500]
    except Exception:
        return False
    return False


# ── Сверка ─────────────────────────────────────────────────────────────────────

def reconcile(conn: sqlite3.Connection, akt: dict) -> dict:
    """Сравнивает накладные из акта с БД. Возвращает diff."""
    # Найти поставщика в БД
    supplier_name = akt.get('supplier') or ''
    sup_row = None
    if supplier_name:
        # Пробуем точный матч, потом fuzzy
        sup_row = conn.execute(
            'SELECT id, name FROM suppliers WHERE name = ?', (supplier_name,),
        ).fetchone()
        if not sup_row:
            # Core-name match (strip ИП/ТОО, кавычки)
            core = re.sub(r'^(ТОО|ИП|АО|ООО|ЧП|КХ)\s*["«]?', '', supplier_name, flags=re.IGNORECASE)
            core = core.strip().strip('"«»').split()[0] if core.strip() else ''
            if core and len(core) > 3:
                sup_row = conn.execute(
                    "SELECT id, name FROM suppliers WHERE name LIKE ?",
                    (f'%{core}%',),
                ).fetchone()

    if not sup_row:
        return {
            'error': f'Поставщик «{supplier_name}» не найден в базе',
            'supplier_lookup': supplier_name,
            **akt,
        }

    sup_id, sup_real_name = sup_row[0], sup_row[1]

    # Накладные из БД за период акта
    pfrom = akt.get('period_from') or '1970-01-01'
    pto   = akt.get('period_to')   or '2999-12-31'
    db_invs = conn.execute(
        """SELECT invoice_id, invoice_date, total_amount
           FROM invoices WHERE supplier_id = ? AND invoice_date BETWEEN ? AND ?""",
        (sup_id, pfrom, pto),
    ).fetchall()
    db_by_id = {str(int(r[0])) if r[0].isdigit() else str(r[0]): {'date': r[1], 'amount': r[2]} for r in db_invs}

    matched, amount_diff, missing_from_db, renumbered = [], [], [], []
    akt_ids_used = set()
    db_ids_used  = set()

    # Pass 1: точный матч по invoice_id
    for a in akt['invoices']:
        aid = a['invoice_id']
        if aid in db_by_id and aid not in db_ids_used:
            akt_ids_used.add(aid); db_ids_used.add(aid)
            d = db_by_id[aid]
            diff = abs((d['amount'] or 0) - a['amount'])
            if diff > 1.0:
                amount_diff.append({
                    'invoice_id': aid, 'date': a['date'],
                    'akt_amount': a['amount'], 'db_amount': d['amount'],
                    'diff': (a['amount'] - d['amount']),
                })
            matched.append({'invoice_id': aid, 'amount': a['amount']})

    # Pass 2: для несматченных — матч по (дата + сумма в пределах 1 ₸)
    # Это покрывает случай когда у поставщика и у нас разная нумерация.
    for a in akt['invoices']:
        if a['invoice_id'] in akt_ids_used: continue
        akt_date = a['date']; akt_amt = a['amount']
        best = None
        for dk, d in db_by_id.items():
            if dk in db_ids_used: continue
            if d['date'] == akt_date and abs((d['amount'] or 0) - akt_amt) <= 1.0:
                best = dk; break
        if best:
            akt_ids_used.add(a['invoice_id']); db_ids_used.add(best)
            renumbered.append({
                'akt_id': a['invoice_id'], 'db_id': best,
                'date': akt_date, 'amount': akt_amt,
            })

    # Что осталось в акте и не нашлось — реально отсутствует в БД
    for a in akt['invoices']:
        if a['invoice_id'] not in akt_ids_used:
            missing_from_db.append(a)

    # Что в БД осталось без матча — возможно лишнее или не входит в период акта
    extra_in_db = [
        {'invoice_id': k, 'date': v['date'], 'amount': v['amount']}
        for k, v in db_by_id.items() if k not in db_ids_used
    ]

    return {
        'supplier': sup_real_name, 'supplier_id': sup_id,
        'period_from': pfrom, 'period_to': pto,
        'akt_total': sum(a['amount'] or 0 for a in akt['invoices']),
        'db_total':  sum((v['amount'] or 0) for v in db_by_id.values()),
        'matched':         matched,
        'renumbered':      renumbered,
        'missing_from_db': missing_from_db,
        'extra_in_db':     extra_in_db,
        'amount_diff':     amount_diff,
    }


def format_report(r: dict) -> str:
    """Текстовый отчёт для Телеграма."""
    if 'error' in r:
        return f"❌ {r['error']}\n\nВ акте найдено {len(r.get('invoices', []))} накладных."

    fmt_n = lambda v: f'{int(round(v)):,}'.replace(',', ' ')
    fmt_d = lambda d: d.split('-')[2] + '.' + d.split('-')[1] if d and '-' in d else (d or '—')

    lines = [
        f"📋 Сверка с {r['supplier']}",
        f"Период: {fmt_d(r['period_from'])} — {fmt_d(r['period_to'])}",
        '',
        f"✅ Совпало по номеру: {len(r['matched'])} на {fmt_n(sum(m['amount'] or 0 for m in r['matched']))} ₸",
    ]

    if r.get('renumbered'):
        lines.append(f"🔄 Совпало по дате+сумме (разная нумерация): {len(r['renumbered'])}")
        for m in r['renumbered'][:8]:
            lines.append(f"  • акт №{m['akt_id']} ↔ база №{m['db_id']} ({fmt_d(m['date'])} · {fmt_n(m['amount'])} ₸)")
        if len(r['renumbered']) > 8:
            lines.append(f"  … ещё {len(r['renumbered']) - 8}")

    if r['missing_from_db']:
        lines.append('')
        lines.append(f"⚠️ У поставщика есть, у тебя НЕТ ({len(r['missing_from_db'])}):")
        for m in r['missing_from_db'][:10]:
            lines.append(f"  • №{m['invoice_id']} от {fmt_d(m['date'])} — {fmt_n(m['amount'])} ₸")
        if len(r['missing_from_db']) > 10:
            lines.append(f"  … ещё {len(r['missing_from_db']) - 10}")

    if r['extra_in_db']:
        lines.append('')
        lines.append(f"❌ У тебя есть, у поставщика НЕТ ({len(r['extra_in_db'])}):")
        for m in r['extra_in_db'][:10]:
            lines.append(f"  • №{m['invoice_id']} от {fmt_d(m['date'])} — {fmt_n(m['amount'])} ₸")

    if r['amount_diff']:
        lines.append('')
        lines.append(f"💰 Расхождение по суммам ({len(r['amount_diff'])}):")
        for d in r['amount_diff'][:10]:
            sign = '+' if d['diff'] > 0 else ''
            lines.append(
                f"  • №{d['invoice_id']}: у тебя {fmt_n(d['db_amount'])}, в акте {fmt_n(d['akt_amount'])} "
                f"(Δ {sign}{fmt_n(d['diff'])} ₸)"
            )

    lines.append('')
    lines.append(f"Итого: акт {fmt_n(r['akt_total'])} ₸ · база {fmt_n(r['db_total'])} ₸")

    return '\n'.join(lines)


# ── Утилиты ────────────────────────────────────────────────────────────────────

def _to_iso(ru_date: str) -> str:
    """'01.04.2026' → '2026-04-01'"""
    try:
        dd, mm, yy = ru_date.split('.')
        return f'{yy}-{mm}-{dd}'
    except Exception:
        return ru_date


def _extract_amount_from_row(row, doc_col_idx):
    """Берёт первую числовую ячейку правее колонки с документом."""
    for cell in row[doc_col_idx + 1:]:
        if cell is None: continue
        try:
            n = to_float(str(cell))
            if n > 0:
                return n
        except Exception:
            continue
    return 0.0
